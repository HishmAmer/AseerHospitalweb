from datetime import date, datetime, timedelta
from functools import wraps

import openpyxl
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.views import LoginView, redirect_to_login
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, ProtectedError
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .excel_compare import ExcelFormatError, compare
from .forms import EmployeeForm, LeaveForm
from .models import (
    ActivityLog,
    Department,
    Employee,
    GeneralSpecialty,
    Leave,
    SubSpecialty,
    UserProfile,
    Workplace,
)
SETTING_ITEM_MODELS = {
    'workplace': Workplace,
    'general_specialty': GeneralSpecialty,
    'sub_specialty': SubSpecialty,
    'department': Department,
}


def superuser_required(view):
    """يقصر العرض على مدير النظام.

    ليس user_passes_test(login_url='login'): صفحة الدخول تعيد المستخدم
    المسجَّل دخوله فوراً إلى next، فإرسال مستخدم فرع مسجَّل إليها كان
    ينتج حلقة تحويل لا تنتهي بدل رفض واضح — على كل صفحات مدير النظام
    لا هذه وحدها. هنا يعود إلى لوحة القيادة برسالة يفهمها.
    """

    @wraps(view)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path(), reverse('login'))
        if not request.user.is_superuser:
            messages.error(request, 'هذه الصفحة متاحة لمدير النظام فقط.')
            return redirect('dashboard')
        return view(request, *args, **kwargs)

    return wrapper


def log_activity(user, action_type, description):
    if user.is_authenticated:
        ActivityLog.objects.create(user=user, action_type=action_type, description=description)


def scoped_employees(user, queryset=None):
    """Employees the user may act on. Fails closed: a non-superuser without a
    workplace on their profile sees nothing rather than everything."""
    qs = Employee.objects.all() if queryset is None else queryset
    if user.is_superuser:
        return qs
    profile = getattr(user, 'profile', None)
    workplace = profile.workplace if profile else None
    if workplace is None:
        return qs.none()
    return qs.filter(current_workplace=workplace)


def scoped_leaves(user, queryset=None):
    qs = Leave.objects.all() if queryset is None else queryset
    if user.is_superuser:
        return qs
    profile = getattr(user, 'profile', None)
    workplace = profile.workplace if profile else None
    if workplace is None:
        return qs.none()
    return qs.filter(employee__current_workplace=workplace)


def scope_label(user):
    profile = getattr(user, 'profile', None)
    if profile and profile.workplace:
        return profile.workplace.name
    return 'الإدارة العامة'


class ThrottledLoginView(LoginView):
    """Login view that locks a username+IP pair out after repeated failures."""

    template_name = 'employees/login.html'
    redirect_authenticated_user = True

    def _throttle_key(self, request):
        username = (request.POST.get('username') or '')[:150]
        forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
        ip = forwarded.split(',')[0].strip() or request.META.get('REMOTE_ADDR', '')
        return f'login-attempts:{username.lower()}:{ip}'

    def post(self, request, *args, **kwargs):
        key = self._throttle_key(request)
        if cache.get(key, 0) >= settings.LOGIN_ATTEMPT_LIMIT:
            messages.error(
                request,
                'تم إيقاف المحاولات مؤقتاً بسبب تكرار إدخال بيانات خاطئة. حاول لاحقاً.',
            )
            return self.render_to_response(self.get_context_data(form=self.get_form()))
        return super().post(request, *args, **kwargs)

    def form_valid(self, form):
        cache.delete(self._throttle_key(self.request))
        return super().form_valid(form)

    def form_invalid(self, form):
        key = self._throttle_key(self.request)
        try:
            cache.incr(key)
        except ValueError:
            cache.set(key, 1, settings.LOGIN_ATTEMPT_TIMEOUT)
        return super().form_invalid(form)


@login_required(login_url='login')
def dashboard_view(request):
    today = date.today()
    base_employees = scoped_employees(request.user, Employee.objects.filter(is_deleted=False))
    on_leave = scoped_leaves(
        request.user,
        Leave.objects.filter(
            employee__is_deleted=False,
            employee__status='نشط',
            start_date__lte=today,
            end_date__gte=today,
        ),
    ).count()

    active_count = base_employees.filter(status='نشط').count()

    thirty_days_from_now = today + timedelta(days=30)
    ninety_days_from_now = today + timedelta(days=90)

    expired_employees = base_employees.filter(
        status='نشط', contract_end_date__lt=today
    ).order_by('contract_end_date')
    expiring_soon_employees = base_employees.filter(
        status='نشط',
        contract_end_date__gte=today,
        contract_end_date__lte=thirty_days_from_now,
    ).order_by('contract_end_date')

    expired_classifications = base_employees.filter(
        status='نشط', is_classified='مصنف', classification_expiry_date__lt=today
    ).order_by('classification_expiry_date')
    expiring_soon_classifications = base_employees.filter(
        status='نشط',
        is_classified='مصنف',
        classification_expiry_date__gte=today,
        classification_expiry_date__lte=ninety_days_from_now,
    ).order_by('classification_expiry_date')

    context = {
        'workplace_name': scope_label(request.user),
        'total_employees': active_count,
        'male_count': base_employees.filter(status='نشط', gender='M').count(),
        'female_count': base_employees.filter(status='نشط', gender='F').count(),
        'on_leave_count': on_leave,
        'active_count': active_count,
        'resigned_count': base_employees.filter(status='مستقيل').count(),
        'terminated_count': base_employees.filter(status='طي القيد').count(),
        'delegated_count': base_employees.filter(status='إيفاد').count(),
        'total_all': base_employees.count(),
        'expired_employees': expired_employees,
        'expiring_soon_employees': expiring_soon_employees,
        'expired_count': expired_employees.count(),
        'expiring_soon_count': expiring_soon_employees.count(),
        'resident_count': base_employees.filter(status='نشط', employee_category='مقيم').count(),
        'registrar_count': base_employees.filter(status='نشط', employee_category='نائب').count(),
        'senior_registrar_count': base_employees.filter(
            status='نشط', employee_category='نائب أول'
        ).count(),
        'consultant_count': base_employees.filter(
            status='نشط', employee_category='استشاري'
        ).count(),
        'expired_classifications': expired_classifications,
        'expiring_soon_classifications': expiring_soon_classifications,
        'expired_class_count': expired_classifications.count(),
        'expiring_soon_class_count': expiring_soon_classifications.count(),
        'departments': Department.objects.all(),
    }
    return render(request, 'employees/dashboard.html', context)


@require_POST
@login_required(login_url='login')
def custom_logout(request):
    logout(request)
    return redirect('login')


def filtered_employees(request):
    """Employees visible to the requester, narrowed by the search filters in the query string."""
    employees = scoped_employees(request.user, Employee.objects.filter(is_deleted=False))

    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()
    workplace_filter = request.GET.get('workplace', '').strip()

    if search_query:
        employees = employees.filter(
            Q(full_name__icontains=search_query)
            | Q(national_id__icontains=search_query)
            | Q(employee_number__icontains=search_query)
        )

    if status_filter in dict(Employee.STATUS_CHOICES):
        employees = employees.filter(status=status_filter)

    if workplace_filter.isdigit() and request.user.is_superuser:
        employees = employees.filter(current_workplace__id=int(workplace_filter))

    return employees.select_related('current_workplace', 'current_department').order_by('full_name')


@login_required(login_url='login')
def employee_list(request):
    context = {
        'employees': filtered_employees(request),
        'search_query': request.GET.get('search', '').strip(),
        'status_filter': request.GET.get('status', '').strip(),
        'workplace_filter': request.GET.get('workplace', '').strip(),
        'workplaces': Workplace.objects.all() if request.user.is_superuser else None,
    }
    return render(request, 'employees/employee_list.html', context)


@login_required(login_url='login')
def add_employee(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST, user=request.user)
        if form.is_valid():
            emp = form.save()
            log_activity(request.user, 'إضافة', f'تم إضافة الموظف: {emp.full_name}')
            messages.success(request, 'تم إضافة الموظف بنجاح!')
            return redirect('employee_list')
    else:
        form = EmployeeForm(user=request.user)
    return render(request, 'employees/add_employee.html', {'form': form})


@login_required(login_url='login')
def edit_employee(request, pk):
    emp = get_object_or_404(scoped_employees(request.user), pk=pk)

    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=emp, user=request.user)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'تعديل', f'تم تعديل بيانات الموظف: {emp.full_name}')
            messages.success(request, 'تم تحديث بيانات الموظف بنجاح!')
            return redirect('employee_list')
    else:
        form = EmployeeForm(instance=emp, user=request.user)
    return render(request, 'employees/edit_employee.html', {'form': form, 'employee': emp})


# أطول سبب يُقبل. الحقل نصّي بلا حدّ في قاعدة البيانات، والحدّ هنا
# يمنع إغراق السجل بنصّ ضخم من طلب مُلفَّق.
ARCHIVE_REASON_MAX = 500


@require_POST
@login_required(login_url='login')
def delete_employee(request, pk):
    employee = get_object_or_404(
        scoped_employees(request.user, Employee.objects.filter(is_deleted=False)), pk=pk
    )

    # السبب إلزامي. النافذة في المتصفح تفرضه، لكن الطلب قد يصل بدونها،
    # فالرفض هنا هو ما يضمن ألّا يدخل الأرشيف سجلٌّ بلا سبب.
    reason = ' '.join((request.POST.get('archive_reason') or '').split())
    if not reason:
        messages.error(request, 'يجب كتابة سبب النقل للأرشيف.')
        return redirect('employee_list')
    if len(reason) > ARCHIVE_REASON_MAX:
        messages.error(
            request, f'سبب النقل طويل جداً (الحد {ARCHIVE_REASON_MAX} حرفاً).'
        )
        return redirect('employee_list')

    employee.is_deleted = True
    employee.archive_reason = reason
    employee.save(update_fields=['is_deleted', 'archive_reason', 'updated_at'])
    log_activity(
        request.user, 'تعديل',
        f'تم نقل الموظف للأرشيف: {employee.full_name} — السبب: {reason}',
    )
    messages.success(request, f'تم نقل الموظف {employee.full_name} للأرشيف.')
    return redirect('employee_list')

def excel_safe(value):
    """Neutralise spreadsheet formula injection in exported cells."""
    if isinstance(value, str) and value[:1] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


@login_required(login_url='login')
def export_employees_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "بيانات الموظفين"
    ws.sheet_view.rightToLeft = True

    ws.append([
        'الرقم الوظيفي', 'اسم الموظف', 'رقم الهوية', 'تاريخ الميلاد', 'العمر', 'الجنسية',
        'نوع المنشأة', 'التفرغ', 'المنشأة الحالية', 'القسم', 'فئة الموظف (نوع العقد)',
        'فئة الكادر', 'المسمى الوظيفي', 'حالة الموظف',
        'تاريخ بداية العقد', 'تاريخ نهاية العقد', 'حالة التصنيف', 'تاريخ انتهاء التصنيف',
    ])

    for emp in filtered_employees(request):
        ws.append([excel_safe(value) for value in (
            emp.employee_number or '-',
            emp.full_name,
            emp.national_id or '-',
            emp.dob.strftime('%Y-%m-%d') if emp.dob else '-',
            emp.age if emp.age is not None else '-',
            emp.nationality or '-',
            emp.workplace_type or '-',
            emp.time_type or '-',
            emp.current_workplace.name if emp.current_workplace else '-',
            emp.current_department.name if emp.current_department else '-',
            emp.employee_type or '-',
            emp.employee_category or '-',
            emp.contract_job_title or '-',
            emp.status,
            emp.contract_start_date.strftime('%Y-%m-%d') if emp.contract_start_date else '-',
            emp.contract_end_date.strftime('%Y-%m-%d') if emp.contract_end_date else '-',
            emp.is_classified,
            emp.classification_expiry_date.strftime('%Y-%m-%d') if emp.classification_expiry_date else '-',
        )])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Employees_Data_{datetime.now().strftime("%Y_%m_%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    log_activity(request.user, 'نظام', 'قام بتصدير بيانات الموظفين المفلترة إلى ملف Excel')
    return response


# أقصى حجم لملف المطابقة. الحدّ منفصل عن DATA_UPLOAD_MAX_MEMORY_SIZE ليعطي
# رسالة عربية واضحة بدل استثناء Django العام.
MAX_UPLOAD_BYTES = 5 * 1024 * 1024


@login_required(login_url='login')
@superuser_required
def reconcile_employees(request):
    """يقارن ملف Excel مرفوعاً بسجلات الموظفين ضمن نطاق صلاحية المستخدم.

    العملية للقراءة فقط — لا تُنشئ ولا تُعدّل أي سجل. الملف لا يُخزَّن في
    مجلد الوسائط؛ يبقى في ذاكرة الطلب، وإن تجاوز FILE_UPLOAD_MAX_MEMORY_SIZE
    كتبه Django في ملف مؤقت يحذفه فور انتهاء الطلب.
    """
    context = {
        'report': None,
        'scope_name': scope_label(request.user),
        'generated_at': timezone.localtime(),
    }

    if request.method == 'POST':
        upload = request.FILES.get('excel_file')

        if upload is None:
            messages.error(request, 'يرجى اختيار ملف Excel أولاً.')
            return render(request, 'employees/reconcile.html', context)

        if not upload.name.lower().endswith('.xlsx'):
            messages.error(
                request,
                'الصيغة غير مدعومة. يجب أن يكون الملف بامتداد ‎.xlsx‎ '
                '(من Excel: حفظ باسم ← Excel Workbook).'
            )
            return render(request, 'employees/reconcile.html', context)

        if upload.size > MAX_UPLOAD_BYTES:
            messages.error(
                request,
                f'حجم الملف {upload.size // 1024 // 1024} ميجابايت، '
                f'والحد الأقصى {MAX_UPLOAD_BYTES // 1024 // 1024} ميجابايت.'
            )
            return render(request, 'employees/reconcile.html', context)

        employees = list(
            scoped_employees(request.user, Employee.objects.filter(is_deleted=False))
            .select_related('current_workplace', 'current_department')
            .order_by('full_name')
        )

        try:
            report = compare(upload, employees)
        except ExcelFormatError as exc:
            messages.error(request, str(exc))
            return render(request, 'employees/reconcile.html', context)

        context['report'] = report
        context['file_name'] = upload.name
        log_activity(
            request.user,
            'نظام',
            f'طابق ملف Excel ({upload.name}) مع بيانات الموظفين: '
            f'{len(report["missing_in_system"])} غير مسجّل، '
            f'{len(report["differences"])} به اختلافات'
        )

    return render(request, 'employees/reconcile.html', context)


@login_required(login_url='login')
@superuser_required
def reconcile_template(request):
    """قالب Excel فارغ بترويسات المطابقة، ليملأه المستخدم ويرفعه."""
    from .excel_compare import COMPARED_FIELDS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'قالب المطابقة'
    ws.sheet_view.rightToLeft = True
    ws.append([label for _, label in COMPARED_FIELDS])

    for index in range(1, len(COMPARED_FIELDS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reconcile_Template.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def leave_list(request):
    today = date.today()
    base_query = scoped_leaves(
        request.user, Leave.objects.filter(employee__is_deleted=False)
    ).select_related('employee')
    return render(request, 'employees/leave_list.html', {
        'active_leaves': base_query.filter(end_date__gte=today).order_by('start_date'),
        'history_leaves': base_query.filter(end_date__lt=today).order_by('-end_date'),
    })


def restrict_leave_employee_choices(form, user):
    form.fields['employee'].queryset = scoped_employees(
        user, Employee.objects.filter(is_deleted=False)
    )


@login_required(login_url='login')
def add_leave(request):
    if request.method == 'POST':
        form = LeaveForm(request.POST)
        restrict_leave_employee_choices(form, request.user)
        if form.is_valid():
            leave = form.save()
            log_activity(request.user, 'إضافة', f'تم تسجيل إجازة للموظف: {leave.employee.full_name}')
            messages.success(request, 'تم تسجيل الإجازة بنجاح!')
            return redirect('leave_list')
    else:
        form = LeaveForm()
        restrict_leave_employee_choices(form, request.user)
    return render(request, 'employees/add_leave.html', {'form': form})


@login_required(login_url='login')
def edit_leave(request, pk):
    leave = get_object_or_404(scoped_leaves(request.user), pk=pk)

    if request.method == 'POST':
        form = LeaveForm(request.POST, instance=leave)
        restrict_leave_employee_choices(form, request.user)
        if form.is_valid():
            form.save()
            log_activity(request.user, 'تعديل', f'تم تعديل إجازة الموظف: {leave.employee.full_name}')
            messages.success(request, 'تم تعديل بيانات الإجازة بنجاح!')
            return redirect('leave_list')
    else:
        form = LeaveForm(instance=leave)
        restrict_leave_employee_choices(form, request.user)
    return render(request, 'employees/edit_leave.html', {'form': form, 'leave': leave})


@require_POST
@login_required(login_url='login')
def delete_leave(request, pk):
    leave = get_object_or_404(scoped_leaves(request.user), pk=pk)
    emp_name = leave.employee.full_name
    leave.delete()
    log_activity(request.user, 'حذف', f'تم إلغاء إجازة الموظف: {emp_name}')
    messages.success(request, 'تم حذف الإجازة بنجاح.')
    return redirect('leave_list')


@login_required(login_url='login')
def archived_employee_list(request):
    archived_emps = scoped_employees(
        request.user, Employee.objects.filter(is_deleted=True)
    ).select_related('current_workplace')
    return render(request, 'employees/archive_list.html', {'employees': archived_emps})


@require_POST
@login_required(login_url='login')
def restore_employee(request, pk):
    emp = get_object_or_404(
        scoped_employees(request.user, Employee.objects.filter(is_deleted=True)), pk=pk
    )
    # السبب يُفرَّغ حتى لا يُنسب سبب أرشفة قديم إلى أرشفة لاحقة؛
    # نصّه محفوظ في سجل النشاطات قبل المسح.
    previous_reason = emp.archive_reason
        # السبب يُفرَّغ حتى لا يُنسب سبب أرشفة قديم إلى أرشفة لاحقة؛
    # نصّه محفوظ في سجل النشاطات قبل المسح.
    previous_reason = emp.archive_reason
    emp.is_deleted = False
    emp.archive_reason = None
    emp.save(update_fields=['is_deleted', 'archive_reason', 'updated_at'])
    log_activity(
        request.user, 'استعادة',
        f'تم استعادة الموظف للعمل: {emp.full_name}'
        + (f' — سبب أرشفته كان: {previous_reason}' if previous_reason else ''),
    )
   
    messages.success(request, f'تم استعادة الموظف {emp.full_name} وإعادته للسجل بنجاح!')
    return redirect('archived_employee_list')


@require_POST
@login_required(login_url='login')
@superuser_required
def hard_delete_employee(request, pk):
    emp = get_object_or_404(Employee, pk=pk, is_deleted=True)
    emp_name = emp.full_name
    emp.delete()
    log_activity(request.user, 'حذف', f'تم مسح الموظف نهائياً من قاعدة البيانات: {emp_name}')
    messages.success(request, 'تم حذف الموظف من قاعدة البيانات نهائياً.')
    return redirect('archived_employee_list')


@login_required(login_url='login')
@superuser_required
def user_list(request):
    users = User.objects.all().select_related('profile', 'profile__workplace')
    return render(request, 'employees/user_list.html', {'users': users})


def resolve_account_workplace(request, role, workplace_id):
    """Validate the role/workplace pairing. Returns (workplace, error_message)."""
    if role == 'admin':
        if workplace_id:
            return None, 'خطأ: حساب الإدارة (الأدمن) لا يجب أن يرتبط بمنشأة محددة.'
        return None, None

    if not workplace_id:
        return None, 'إجباري: يجب اختيار المنشأة للحسابات الفرعية.'

    workplace = Workplace.objects.filter(pk=workplace_id).first()
    if workplace is None:
        return None, 'المنشأة المختارة غير موجودة.'
    return workplace, None


@login_required(login_url='login')
@superuser_required
def add_user(request):
    workplaces = Workplace.objects.all()
    if request.method == 'POST':
        username = (request.POST.get('username') or '').strip()
        password = request.POST.get('password') or ''
        first_name = (request.POST.get('first_name') or '').strip()
        role = request.POST.get('role')
        status = request.POST.get('status')

        workplace, error = resolve_account_workplace(request, role, request.POST.get('workplace'))
        if error:
            messages.error(request, error)
            return redirect('add_user')

        if not username:
            messages.error(request, 'اسم المستخدم مطلوب.')
            return redirect('add_user')

        if User.objects.filter(username__iexact=username).exists():
            messages.error(request, 'اسم المستخدم هذا موجود مسبقاً!')
            return redirect('add_user')

        try:
            validate_password(password, User(username=username, first_name=first_name))
        except ValidationError as exc:
            messages.error(request, ' '.join(exc.messages))
            return redirect('add_user')

        is_admin = role == 'admin'
        with transaction.atomic():
            new_user = User.objects.create_user(
                username=username,
                password=password,
                first_name=first_name,
                is_superuser=is_admin,
                is_staff=is_admin,
                is_active=status != 'inactive',
            )
            UserProfile.objects.create(user=new_user, workplace=workplace)

        log_activity(request.user, 'إضافة', f'تم إنشاء مستخدم جديد: {username}')
        messages.success(request, f'تم إنشاء المستخدم ({username}) وربطه بنجاح!')
        return redirect('user_list')
    return render(request, 'employees/add_user.html', {'workplaces': workplaces})


@login_required(login_url='login')
@superuser_required
def edit_user_permissions(request, pk):
    target_user = get_object_or_404(User, pk=pk)
    workplaces = Workplace.objects.all()

    if request.method == 'POST':
        role = request.POST.get('role')
        status = request.POST.get('status')
        is_self = target_user == request.user

        # Guard against an administrator locking themselves out of the system.
        if is_self and status == 'inactive':
            messages.error(request, 'عملية مرفوضة: لا يمكنك تعطيل حسابك الشخصي أثناء استخدامه!')
            return redirect('user_list')
        if is_self and role != 'admin':
            messages.error(request, 'عملية مرفوضة: لا يمكنك سحب صلاحيات الإدارة من حسابك الشخصي!')
            return redirect('user_list')

        workplace, error = resolve_account_workplace(request, role, request.POST.get('workplace'))
        if error:
            messages.error(request, error)
            return redirect('edit_user', pk=target_user.pk)

        new_username = (request.POST.get('username') or '').strip()
        if new_username and new_username.lower() != target_user.username.lower():
            if User.objects.filter(username__iexact=new_username).exists():
                messages.error(request, 'اسم المستخدم هذا مسجل بالفعل لمستخدم آخر!')
                return redirect('edit_user', pk=target_user.pk)
            target_user.username = new_username

        new_password = request.POST.get('password')
        if new_password:
            try:
                validate_password(new_password, target_user)
            except ValidationError as exc:
                messages.error(request, ' '.join(exc.messages))
                return redirect('edit_user', pk=target_user.pk)
            target_user.set_password(new_password)

        is_admin = role == 'admin'
        target_user.is_superuser = is_admin
        target_user.is_staff = is_admin
        target_user.is_active = status == 'active'

        with transaction.atomic():
            target_user.save()
            UserProfile.objects.update_or_create(
                user=target_user, defaults={'workplace': workplace}
            )

        log_activity(
            request.user, 'تعديل', f'تم تعديل بيانات وصلاحيات المستخدم: {target_user.username}'
        )
        messages.success(request, f'تم تحديث بيانات المستخدم ({target_user.username}) بنجاح!')
        return redirect('user_list')

    return render(
        request, 'employees/edit_user.html', {'target_user': target_user, 'workplaces': workplaces}
    )


@login_required(login_url='login')
@superuser_required
def system_settings(request):
    context = {
        'workplaces': Workplace.objects.all(),
        'departments': Department.objects.all(),
        'general_specialties': GeneralSpecialty.objects.all(),
        'sub_specialties': SubSpecialty.objects.all(),
    }
    return render(request, 'employees/settings.html', context)


@require_POST
@login_required(login_url='login')
@superuser_required
def add_setting_item(request, item_type):
    model = SETTING_ITEM_MODELS.get(item_type)
    name = (request.POST.get('name') or '').strip()
    if model and name:
        model.objects.create(name=name)
        log_activity(request.user, 'نظام', f'تم إضافة {name} إلى إعدادات النظام')
        messages.success(request, 'تمت الإضافة بنجاح!')
    else:
        messages.error(request, 'تعذر تنفيذ الإضافة: البيانات المدخلة غير صحيحة.')
    return redirect('system_settings')


@require_POST
@login_required(login_url='login')
@superuser_required
def delete_setting_item(request, item_type, pk):
    model = SETTING_ITEM_MODELS.get(item_type)
    if model is None:
        messages.error(request, 'نوع العنصر غير معروف.')
        return redirect('system_settings')

    obj = get_object_or_404(model, pk=pk)
    name = obj.name
    try:
        obj.delete()
    except ProtectedError:
        messages.error(request, 'لا يمكن الحذف! هذا العنصر مرتبط ببيانات موظفين حاليين.')
        return redirect('system_settings')

    log_activity(request.user, 'نظام', f'تم حذف {name} من إعدادات النظام')
    messages.success(request, 'تم الحذف بنجاح.')
    return redirect('system_settings')


@login_required(login_url='login')
@superuser_required
def activity_log_list(request):
    thirty_days_ago = timezone.now() - timedelta(days=30)
    logs = (
        ActivityLog.objects.filter(timestamp__gte=thirty_days_ago)
        .select_related('user')
        .order_by('-timestamp')
    )
    return render(request, 'employees/activity_log.html', {'logs': logs})