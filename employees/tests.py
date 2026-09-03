import re
import sys
from datetime import date, timedelta
from unittest import mock
from django.contrib.auth.models import User
from django.core.cache import cache
from django.test import SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone

from .forms import EmployeeForm, latest_acceptable_dob
from .templatetags.arabic_time import days_ago
from .models import ActivityLog, Employee, Leave, Nationality, UserProfile, Workplace


class SecurityTestCase(TestCase):
    def setUp(self):
        cache.clear()
        self.hospital_a = Workplace.objects.create(name='مستشفى أ')
        self.hospital_b = Workplace.objects.create(name='مستشفى ب')

        self.admin = User.objects.create_user(
            username='admin', password='Str0ngAdminPass!', is_superuser=True, is_staff=True
        )
        UserProfile.objects.create(user=self.admin, workplace=None)

        # الجنسيات تُزرع في هجرة بيانات، فهي موجودة في قاعدة الاختبار.
        self.saudi = Nationality.objects.get(name='سعودي')

        self.branch_user = User.objects.create_user(username='branch_a', password='Str0ngPass!2024')
        UserProfile.objects.create(user=self.branch_user, workplace=self.hospital_a)

        self.own_employee = Employee.objects.create(
            full_name='موظف أ', gender='M', current_workplace=self.hospital_a
        )
        self.other_employee = Employee.objects.create(
            full_name='موظف ب', gender='F', current_workplace=self.hospital_b
        )


class AuthenticationRequiredTests(SecurityTestCase):
    def test_system_settings_rejects_anonymous(self):
        response = self.client.get(reverse('system_settings'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])

    def test_system_settings_rejects_non_superuser(self):
        # الوجهة لوحة القيادة لا صفحة الدخول: إعادة مستخدم مسجَّل إلى
        # الدخول تُنتج حلقة تحويل، لأن تلك الصفحة تعيده إلى next فوراً.
        self.client.force_login(self.branch_user)
        response = self.client.get(reverse('system_settings'))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('dashboard'))

    def test_system_settings_allows_superuser(self):
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse('system_settings')).status_code, 200)


class DestructiveActionsRequirePostTests(SecurityTestCase):
    def test_get_requests_do_not_mutate_data(self):
        self.client.force_login(self.admin)
        leave = Leave.objects.create(
            employee=self.own_employee,
            leave_type='سنوية',
            start_date='2026-01-01',
            end_date='2026-01-05',
        )
        archived = Employee.objects.create(
            full_name='مؤرشف', gender='M', current_workplace=self.hospital_a, is_deleted=True
        )

        get_only_urls = [
            reverse('delete_employee', args=[self.own_employee.pk]),
            reverse('delete_leave', args=[leave.pk]),
            reverse('restore_employee', args=[archived.pk]),
            reverse('hard_delete_employee', args=[archived.pk]),
            reverse('delete_setting_item', args=['workplace', self.hospital_b.pk]),
            reverse('add_setting_item', args=['workplace']),
            reverse('logout'),
        ]
        for url in get_only_urls:
            with self.subTest(url=url):
                self.assertEqual(self.client.get(url).status_code, 405)

        self.own_employee.refresh_from_db()
        self.assertFalse(self.own_employee.is_deleted)
        self.assertTrue(Leave.objects.filter(pk=leave.pk).exists())
        self.assertTrue(Workplace.objects.filter(pk=self.hospital_b.pk).exists())
        archived.refresh_from_db()
        self.assertTrue(archived.is_deleted)


class WorkplaceScopingTests(SecurityTestCase):
    def test_cannot_view_other_workplace_employee(self):
        self.client.force_login(self.branch_user)
        response = self.client.get(reverse('edit_employee', args=[self.other_employee.pk]))
        self.assertEqual(response.status_code, 404)

    def test_cannot_delete_other_workplace_employee(self):
        self.client.force_login(self.branch_user)
        response = self.client.post(reverse('delete_employee', args=[self.other_employee.pk]))
        self.assertEqual(response.status_code, 404)
        self.other_employee.refresh_from_db()
        self.assertFalse(self.other_employee.is_deleted)

    def test_cannot_delete_other_workplace_leave(self):
        leave = Leave.objects.create(
            employee=self.other_employee,
            leave_type='سنوية',
            start_date='2026-01-01',
            end_date='2026-01-05',
        )
        self.client.force_login(self.branch_user)
        self.assertEqual(
            self.client.post(reverse('delete_leave', args=[leave.pk])).status_code, 404
        )
        self.assertTrue(Leave.objects.filter(pk=leave.pk).exists())

    def test_cannot_file_leave_for_other_workplace_employee(self):
        self.client.force_login(self.branch_user)
        self.client.post(reverse('add_leave'), {
            'employee': self.other_employee.pk,
            'leave_type': 'سنوية',
            'start_date': '2026-01-01',
            'end_date': '2026-01-05',
            'notes': '',
        })
        self.assertFalse(Leave.objects.filter(employee=self.other_employee).exists())

    def test_export_excludes_other_workplace_employees(self):
        self.client.force_login(self.branch_user)
        response = self.client.get(reverse('export_employees_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('موظف ب'.encode(), response.content)

    def test_user_without_profile_sees_nothing(self):
        orphan = User.objects.create_user(username='orphan', password='Str0ngPass!2024')
        self.client.force_login(orphan)
        response = self.client.get(reverse('employee_list'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['employees']), [])

    def test_hard_delete_restricted_to_superuser(self):
        archived = Employee.objects.create(
            full_name='مؤرشف أ', gender='M', current_workplace=self.hospital_a, is_deleted=True
        )
        self.client.force_login(self.branch_user)
        self.client.post(reverse('hard_delete_employee', args=[archived.pk]))
        self.assertTrue(Employee.objects.filter(pk=archived.pk).exists())


class AccountManagementTests(SecurityTestCase):
    def test_weak_password_is_rejected(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('add_user'), {
            'username': 'weakling',
            'password': '123456',
            'first_name': 'Test',
            'role': 'branch',
            'status': 'active',
            'workplace': self.hospital_a.pk,
        })
        self.assertFalse(User.objects.filter(username='weakling').exists())

    def test_strong_password_is_accepted(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('add_user'), {
            'username': 'newuser',
            'password': 'Str0ngPass!2024',
            'first_name': 'Test',
            'role': 'branch',
            'status': 'active',
            'workplace': self.hospital_a.pk,
        })
        self.assertTrue(User.objects.filter(username='newuser').exists())

    def test_unknown_workplace_does_not_error(self):
        self.client.force_login(self.admin)
        response = self.client.post(reverse('add_user'), {
            'username': 'ghost',
            'password': 'Str0ngPass!2024',
            'role': 'branch',
            'status': 'active',
            'workplace': '999999',
        })
        self.assertEqual(response.status_code, 302)
        self.assertFalse(User.objects.filter(username='ghost').exists())

    def test_admin_cannot_revoke_own_admin_rights(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('edit_user', args=[self.admin.pk]), {
            'username': 'admin',
            'role': 'branch',
            'status': 'active',
            'workplace': self.hospital_a.pk,
        })
        self.admin.refresh_from_db()
        self.assertTrue(self.admin.is_superuser)

    def test_non_superuser_cannot_manage_users(self):
        self.client.force_login(self.branch_user)
        response = self.client.get(reverse('user_list'))
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('dashboard'))


class LoginThrottleTests(SecurityTestCase):
    def test_repeated_failures_lock_the_account_out(self):
        url = reverse('login')
        for _ in range(10):
            self.client.post(url, {'username': 'branch_a', 'password': 'wrong'})

        response = self.client.post(url, {'username': 'branch_a', 'password': 'Str0ngPass!2024'})
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.wsgi_request.user.is_authenticated)

    def test_successful_login_clears_the_counter(self):
        url = reverse('login')
        self.client.post(url, {'username': 'branch_a', 'password': 'wrong'})
        response = self.client.post(url, {'username': 'branch_a', 'password': 'Str0ngPass!2024'})
        self.assertEqual(response.status_code, 302)


class ExcelExportTests(SecurityTestCase):
    def test_formula_is_neutralised(self):
        from .views import excel_safe

        self.assertEqual(excel_safe('=cmd|calc'), "'=cmd|calc")
        self.assertEqual(excel_safe('+1+1'), "'+1+1")
        self.assertEqual(excel_safe('محمد'), 'محمد')
        self.assertEqual(excel_safe(42), 42)
        # حرف واحد ليس صيغة؛ الشرطة النائبة عن الخانة الفارغة تخرج كما هي.
        self.assertEqual(excel_safe('-'), '-')
        self.assertEqual(excel_safe('='), '=')
        self.assertEqual(excel_safe('-1+1'), "'-1+1")

    def test_export_includes_specialties(self):
        """التخصص العام والدقيق عمودان في ملف التصدير."""
        import io

        import openpyxl

        from .models import GeneralSpecialty, SubSpecialty

        employee = Employee.objects.get(pk=self.own_employee.pk)
        employee.general_specialty = GeneralSpecialty.objects.create(name='باطنة')
        employee.has_sub_specialty = 'نعم'
        employee.sub_specialty = SubSpecialty.objects.create(name='أمراض الكلى')
        employee.save()

        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.get(reverse('export_employees_excel'))
        self.assertEqual(response.status_code, 200)

        sheet = openpyxl.load_workbook(io.BytesIO(response.content)).active
        header = [cell.value for cell in sheet[1]]
        self.assertIn('التخصص العام', header)
        self.assertIn('التخصص الدقيق', header)

        row = next(
            r for r in sheet.iter_rows(min_row=2, values_only=True)
            if r[header.index('اسم الموظف')] == employee.full_name
        )
        self.assertEqual(row[header.index('التخصص العام')], 'باطنة')
        self.assertEqual(row[header.index('التخصص الدقيق')], 'أمراض الكلى')

    def test_export_includes_classification_number(self):
        """رقم التصنيف عمود في ملف التصدير، وشرطة لغير المصنَّف."""
        import io

        import openpyxl

        classified = Employee.objects.create(
            full_name='موظف مصنف', gender='M', current_workplace=self.hospital_a,
            is_classified='مصنف', classification_number='SCFHS-4521',
        )

        self.client.login(username='admin', password='Str0ngAdminPass!')
        sheet = openpyxl.load_workbook(
            io.BytesIO(self.client.get(reverse('export_employees_excel')).content)
        ).active
        header = [cell.value for cell in sheet[1]]
        self.assertIn('رقم التصنيف', header)

        rows = {
            r[header.index('اسم الموظف')]: r
            for r in sheet.iter_rows(min_row=2, values_only=True)
        }
        self.assertEqual(
            rows[classified.full_name][header.index('رقم التصنيف')], 'SCFHS-4521'
        )
        self.assertEqual(
            rows[self.own_employee.full_name][header.index('رقم التصنيف')], '-'
        )

    def test_export_includes_admin_assignment(self):
        """التكليف الإداري يخرج نعم/لا لا True/False."""
        import io

        import openpyxl

        assigned = Employee.objects.create(
            full_name='موظف مكلف', gender='F', current_workplace=self.hospital_a,
            is_admin_assigned=True,
        )

        self.client.login(username='admin', password='Str0ngAdminPass!')
        sheet = openpyxl.load_workbook(
            io.BytesIO(self.client.get(reverse('export_employees_excel')).content)
        ).active
        header = [cell.value for cell in sheet[1]]
        self.assertIn('مكلف بعمل إداري', header)

        rows = {
            r[header.index('اسم الموظف')]: r
            for r in sheet.iter_rows(min_row=2, values_only=True)
        }
        column = header.index('مكلف بعمل إداري')
        self.assertEqual(rows[assigned.full_name][column], 'نعم')
        self.assertEqual(rows[self.own_employee.full_name][column], 'لا')

    def test_export_shows_dash_when_no_specialty(self):
        """موظف بلا تخصص يخرج بشرطة، لا بخلية فارغة أو خطأ."""
        import io

        import openpyxl

        self.client.login(username='admin', password='Str0ngAdminPass!')
        sheet = openpyxl.load_workbook(
            io.BytesIO(self.client.get(reverse('export_employees_excel')).content)
        ).active
        header = [cell.value for cell in sheet[1]]
        row = next(
            r for r in sheet.iter_rows(min_row=2, values_only=True)
            if r[header.index('اسم الموظف')] == self.own_employee.full_name
        )
        self.assertEqual(row[header.index('التخصص العام')], '-')
        self.assertEqual(row[header.index('التخصص الدقيق')], '-')


class HostConfigurationTests(SimpleTestCase):
    """Guards the deployment contract: a build must not need a hostname it
    cannot know, and a serving process must never run without one."""

    def test_pasted_urls_reduce_to_a_bare_hostname(self):
        from core.settings import as_hostname

        for value in ('https://app.onrender.com/', 'app.onrender.com:443', 'App.onrender.com.'):
            self.assertEqual(as_hostname(value), 'app.onrender.com')

    def test_ipv6_literal_keeps_its_brackets(self):
        from core.settings import as_hostname

        self.assertEqual(as_hostname('[::1]:8000'), '[::1]')

    def test_csrf_origin_gains_a_scheme_and_is_lowercased(self):
        from core.settings import as_origin

        self.assertEqual(as_origin('App.onrender.com', 'https'), 'https://app.onrender.com')
        self.assertEqual(as_origin('http://10.0.0.15/', 'https'), 'http://10.0.0.15')

    def test_build_commands_do_not_require_allowed_hosts(self):
        from core.settings import will_serve_requests

        for argv in (['manage.py', 'collectstatic'], ['manage.py', 'migrate']):
            with mock.patch.object(sys, 'argv', argv):
                self.assertFalse(will_serve_requests())

    def test_serving_processes_require_allowed_hosts(self):
        from core.settings import will_serve_requests

        with mock.patch.object(sys, 'argv', ['manage.py', 'runserver']):
            self.assertTrue(will_serve_requests())
        # gunicorn imports settings without manage.py on argv.
        with mock.patch.object(sys, 'argv', ['/usr/bin/gunicorn', 'core.wsgi']):
            self.assertTrue(will_serve_requests())


def build_workbook(rows, headers=None):
    """ملف Excel في الذاكرة، بترويسات التصدير الافتراضية."""
    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers or ['اسم الموظف', 'رقم الهوية', 'الرقم الوظيفي', 'حالة الموظف'])
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    buffer.name = 'test.xlsx'
    return buffer


class ExcelReconcileTests(SecurityTestCase):
    def setUp(self):
        super().setUp()
        self.own_employee.national_id = '1010101010'
        self.own_employee.employee_number = 'EMP-1'
        self.own_employee.save()

        self.other_employee.national_id = '2020202020'
        self.other_employee.employee_number = 'EMP-2'
        self.other_employee.save()

        self.url = reverse('reconcile_employees')

    # المطابقة صارت لمدير النظام وحده، فهو المستخدم الافتراضي هنا.
    def post_file(self, buffer, user='admin', password='Str0ngAdminPass!'):
        self.client.login(username=user, password=password)
        return self.client.post(self.url, {'excel_file': buffer})

    def test_page_requires_login(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])

    def test_row_absent_from_system_is_reported(self):
        buffer = build_workbook([['موظف جديد', '3030303030', 'EMP-9', 'نشط']])
        report = self.post_file(buffer).context['report']

        self.assertEqual(len(report['missing_in_system']), 1)
        self.assertEqual(report['missing_in_system'][0]['national_id'], '3030303030')

    def test_employee_absent_from_file_is_reported(self):
        """موظف في النظام ولا صفَّ له في الملف يظهر في «الناقص من الملف»."""
        buffer = build_workbook([['موظف أ', '1010101010', 'EMP-1', 'نشط']])
        report = self.post_file(buffer).context['report']

        self.assertEqual(report['matched_count'], 1)
        self.assertEqual(
            [e.full_name for e in report['missing_in_file']], ['موظف ب']
        )

    def test_differing_value_is_reported(self):
        buffer = build_workbook([['موظف أ', '1010101010', 'EMP-1', 'مستقيل']])
        report = self.post_file(buffer).context['report']

        self.assertEqual(len(report['differences']), 1)
        changed = report['differences'][0]['fields']
        self.assertEqual([f['label'] for f in changed], ['حالة الموظف'])
        self.assertEqual(changed[0]['system'], 'نشط')
        self.assertEqual(changed[0]['file'], 'مستقيل')

    def test_branch_user_is_denied(self):
        """المطابقة لمدير النظام وحده: مستخدم الفرع يُصرف عن الصفحة."""
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('dashboard'))

    def test_branch_user_cannot_post_a_file_either(self):
        """المنع على الطلب نفسه لا على الرابط في القائمة فقط."""
        buffer = build_workbook([['موظف أ', '1010101010', 'EMP-1', 'نشط']])
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        response = self.client.post(self.url, {'excel_file': buffer})

        self.assertEqual(response.status_code, 302)

    def test_branch_user_is_denied_the_template(self):
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        response = self.client.get(reverse('reconcile_template'))

        self.assertEqual(response.status_code, 302)

    def test_admin_sees_every_workplace(self):
        """مدير النظام غير محصور بمنشأة، فالتقرير يشمل الجميع."""
        buffer = build_workbook([['موظف ب', '2020202020', 'EMP-2', 'نشط']])
        report = self.post_file(buffer).context['report']

        self.assertEqual(report['total_system'], 2)
        self.assertEqual(report['matched_count'], 1)

    def test_arabic_digits_and_numeric_cells_match(self):
        """رقم الهوية كرقم أو بأرقام عربية يطابق النص المخزَّن."""
        buffer = build_workbook([['موظف أ', 1010101010, 'EMP-1', 'نشط']])
        self.assertEqual(self.post_file(buffer).context['report']['matched_count'], 1)

        buffer = build_workbook([['موظف أ', '١٠١٠١٠١٠١٠', 'EMP-1', 'نشط']])
        self.assertEqual(self.post_file(buffer).context['report']['matched_count'], 1)

    def test_duplicate_row_is_flagged_not_double_counted(self):
        buffer = build_workbook([
            ['موظف أ', '1010101010', 'EMP-1', 'نشط'],
            ['موظف أ', '1010101010', 'EMP-1', 'نشط'],
        ])
        report = self.post_file(buffer).context['report']

        self.assertEqual(report['matched_count'], 1)
        self.assertEqual(len(report['unreadable_rows']), 1)
        self.assertIn('مكرر', report['unreadable_rows'][0]['reason'])

    def test_row_without_any_identifier_is_flagged(self):
        buffer = build_workbook([['بلا معرّف', None, None, 'نشط']])
        report = self.post_file(buffer).context['report']

        self.assertEqual(len(report['unreadable_rows']), 1)
        self.assertEqual(report['missing_in_system'], [])

    def test_file_without_identifier_column_is_rejected(self):
        buffer = build_workbook([['موظف أ', 'نشط']], headers=['اسم الموظف', 'حالة الموظف'])
        response = self.post_file(buffer)

        self.assertIsNone(response.context['report'])
        self.assertContains(response, 'رقم الهوية', status_code=200)

    def test_non_xlsx_upload_is_rejected(self):
        import io

        fake = io.BytesIO(b'name,id\nfoo,1\n')
        fake.name = 'employees.csv'
        response = self.post_file(fake)

        self.assertIsNone(response.context['report'])

    def test_comparison_never_writes_to_the_database(self):
        before = list(Employee.objects.values_list('full_name', 'status', 'national_id'))
        buffer = build_workbook([
            ['اسم مختلف تماماً', '1010101010', 'EMP-1', 'مستقيل'],
            ['موظف جديد', '9090909090', 'EMP-99', 'نشط'],
        ])
        self.post_file(buffer)

        self.assertEqual(before, list(Employee.objects.values_list('full_name', 'status', 'national_id')))

    def test_template_download_is_an_xlsx(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.get(reverse('reconcile_template'))

        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])


class OtherWorkplaceTests(SecurityTestCase):
    """خيار «أخرى» في المنشأة الحالية: يُنشئ منشأة حقيقية، ولمدير النظام وحده."""

    def base_payload(self, **overrides):
        payload = {
            'full_name': 'موظف جديد',
            'gender': 'M',
            'nationality': self.saudi.pk,
            'status': 'نشط',
            'is_classified': 'غير مصنف',
            'has_sub_specialty': 'لا',
            'current_workplace': '__other__',
            'new_workplace_name': 'مستشفى خميس مشيط العام',
        }
        payload.update(overrides)
        return payload

    def test_admin_creates_the_workplace_by_name(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(reverse('add_employee'), self.base_payload())

        self.assertEqual(response.status_code, 302)
        workplace = Workplace.objects.get(name='مستشفى خميس مشيط العام')
        employee = Employee.objects.get(full_name='موظف جديد')
        self.assertEqual(employee.current_workplace, workplace)

    def test_existing_workplace_is_reused_not_duplicated(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        self.client.post(self.base_payload_url(), self.base_payload(
            new_workplace_name='  مستشفى   أ  '))          # مسافات زائدة

        self.assertEqual(Workplace.objects.filter(name__iexact='مستشفى أ').count(), 1)
        self.assertEqual(Employee.objects.get(full_name='موظف جديد').current_workplace,
                         self.hospital_a)

    def base_payload_url(self):
        return reverse('add_employee')

    def test_blank_name_is_rejected(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(
            reverse('add_employee'), self.base_payload(new_workplace_name='   ')
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Employee.objects.filter(full_name='موظف جديد').exists())
        self.assertIn('new_workplace_name', response.context['form'].errors)

    def test_nothing_is_created_when_another_field_fails(self):
        """المنشأة تُنشأ عند الحفظ لا أثناء التحقق، فلا تبقى معلّقة عند الفشل."""
        self.client.login(username='admin', password='Str0ngAdminPass!')
        before = Workplace.objects.count()
        response = self.client.post(reverse('add_employee'), self.base_payload(
            full_name='',                                   # حقل مطلوب فارغ
            new_workplace_name='منشأة لا يجب أن تُنشأ',
        ))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Workplace.objects.count(), before)
        self.assertFalse(Workplace.objects.filter(name='منشأة لا يجب أن تُنشأ').exists())

    def test_branch_user_cannot_use_other(self):
        """مستخدم الفرع محصور بمنشأته — وإلا أنشأ موظفاً يختفي عنه فوراً."""
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        before = Workplace.objects.count()
        response = self.client.post(reverse('add_employee'), self.base_payload())

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Workplace.objects.count(), before)
        self.assertFalse(Employee.objects.filter(full_name='موظف جديد').exists())

    def test_option_is_absent_for_a_branch_user(self):
        # النص يرد في سكربت الصفحة لكل المستخدمين؛ المهم غيابه كخيار في القائمة.
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        self.assertNotContains(
            self.client.get(reverse('add_employee')), 'value="__other__"'
        )

    def test_option_is_offered_to_the_admin(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        self.assertContains(
            self.client.get(reverse('add_employee')), 'value="__other__"'
        )

    def test_normal_selection_still_works(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        before = Workplace.objects.count()
        self.client.post(reverse('add_employee'), self.base_payload(
            current_workplace=str(self.hospital_b.pk),
            new_workplace_name='لا يجب استخدامه',
        ))

        self.assertEqual(Workplace.objects.count(), before)
        self.assertEqual(Employee.objects.get(full_name='موظف جديد').current_workplace,
                         self.hospital_b)


class OtherWorkplaceTypeTests(SecurityTestCase):
    """خيار «أخرى» في نوع المنشأة: النص المكتوب يُخزَّن في الحقل نفسه."""

    def payload(self, **overrides):
        data = {
            'full_name': 'موظف نوع مخصص',
            'gender': 'M',
            'nationality': self.saudi.pk,
            'status': 'نشط',
            'is_classified': 'غير مصنف',
            'has_sub_specialty': 'لا',
            'current_workplace': str(self.hospital_a.pk),
            'workplace_type': '__other_type__',
            'other_workplace_type': 'قطاع خاص',
        }
        data.update(overrides)
        return data

    def test_typed_type_is_stored_on_the_employee(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(reverse('add_employee'), self.payload())

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            Employee.objects.get(full_name='موظف نوع مخصص').workplace_type, 'قطاع خاص'
        )

    def test_whitespace_is_collapsed(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        self.client.post(reverse('add_employee'),
                         self.payload(other_workplace_type='  قطاع   خاص  '))

        self.assertEqual(
            Employee.objects.get(full_name='موظف نوع مخصص').workplace_type, 'قطاع خاص'
        )

    def test_blank_typed_type_is_rejected(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(reverse('add_employee'),
                                    self.payload(other_workplace_type='  '))

        self.assertEqual(response.status_code, 200)
        self.assertIn('other_workplace_type', response.context['form'].errors)
        self.assertFalse(Employee.objects.filter(full_name='موظف نوع مخصص').exists())

    def test_listed_type_still_saves_normally(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        self.client.post(reverse('add_employee'), self.payload(
            workplace_type='مدن طبية', time_type='كلي',
            other_workplace_type='يجب تجاهله',
        ))

        self.assertEqual(
            Employee.objects.get(full_name='موظف نوع مخصص').workplace_type, 'مدن طبية'
        )

    def test_editing_a_custom_type_preselects_other(self):
        """بدون هذا يظهر الحقل فارغاً فتضيع القيمة المحفوظة عند أول حفظ."""
        employee = Employee.objects.create(
            full_name='موظف قائم', gender='M',
            current_workplace=self.hospital_a, workplace_type='قطاع خاص',
        )
        form = EmployeeForm(instance=employee, user=self.admin)

        self.assertEqual(form.initial['workplace_type'], '__other_type__')
        self.assertEqual(form.initial['other_workplace_type'], 'قطاع خاص')

    def test_editing_a_listed_type_is_untouched(self):
        employee = Employee.objects.create(
            full_name='موظف قائم', gender='M',
            current_workplace=self.hospital_a, workplace_type='مدن طبية',
        )
        form = EmployeeForm(instance=employee, user=self.admin)

        self.assertNotIn('other_workplace_type', form.initial)

    def test_branch_user_may_type_a_type(self):
        """نوع المنشأة وصف لا يحدّد صلاحية، فلا سبب لحصره بمدير النظام."""
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        response = self.client.post(reverse('add_employee'), self.payload())

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            Employee.objects.get(full_name='موظف نوع مخصص').workplace_type, 'قطاع خاص'
        )


class ContractDatesByEmployeeTypeTests(SecurityTestCase):
    """تواريخ العقد تتبع «فئة الموظف» لا «فئة الموظف (الكادر)».

    خدمة مدنية: لا عقد، فأي تاريخ يُرسل يُمحى.
    غير ذلك:    تاريخ نهاية العقد إلزامي.
    """

    def payload(self, **overrides):
        data = {
            'full_name': 'موظف عقد',
            'gender': 'M',
            'nationality': self.saudi.pk,
            'status': 'نشط',
            'is_classified': 'غير مصنف',
            'has_sub_specialty': 'لا',
            'current_workplace': self.hospital_a.pk,
        }
        data.update(overrides)
        return data

    def test_civil_service_saves_without_contract_dates(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(
            reverse('add_employee'), self.payload(employee_type='خدمة مدنية')
        )

        self.assertEqual(response.status_code, 302)
        employee = Employee.objects.get(full_name='موظف عقد')
        self.assertIsNone(employee.contract_end_date)

    def test_civil_service_discards_dates_sent_anyway(self):
        """الإخفاء في المتصفح لا يمنع إرسال الحقلين يدوياً."""
        self.client.login(username='admin', password='Str0ngAdminPass!')
        self.client.post(reverse('add_employee'), self.payload(
            employee_type='خدمة مدنية',
            contract_start_date='2026-01-01',
            contract_end_date='2026-12-31',
        ))

        employee = Employee.objects.get(full_name='موظف عقد')
        self.assertIsNone(employee.contract_start_date)
        self.assertIsNone(employee.contract_end_date)

    def test_other_type_requires_an_end_date(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(
            reverse('add_employee'), self.payload(employee_type='تشغيل ذاتي')
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(Employee.objects.filter(full_name='موظف عقد').exists())
        self.assertIn('contract_end_date', response.context['form'].errors)

    def test_other_type_saves_with_an_end_date(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(reverse('add_employee'), self.payload(
            employee_type='تشغيل ذاتي',
            contract_start_date='2026-01-01',
            contract_end_date='2026-12-31',
        ))

        self.assertEqual(response.status_code, 302)
        employee = Employee.objects.get(full_name='موظف عقد')
        self.assertEqual(str(employee.contract_end_date), '2026-12-31')

    def test_blank_type_requires_nothing(self):
        """الحقل نفسه اختياري، فلا يُلزم من تركه فارغاً بتاريخ عقد."""
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(reverse('add_employee'), self.payload())

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(Employee.objects.get(full_name='موظف عقد').contract_end_date)

    def test_end_before_start_is_still_rejected(self):
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(reverse('add_employee'), self.payload(
            employee_type='تشغيل ذاتي',
            contract_start_date='2026-12-31',
            contract_end_date='2026-01-01',
        ))

        self.assertEqual(response.status_code, 200)
        self.assertIn('contract_end_date', response.context['form'].errors)

    def test_switching_to_civil_service_clears_stored_dates(self):
        """موظف بعقد قائم غُيِّرت فئته: لا تبقى تواريخ عقد معلّقة."""
        employee = Employee.objects.create(
            full_name='موظف قائم', gender='M', current_workplace=self.hospital_a,
            employee_type='تشغيل ذاتي',
            contract_start_date='2025-01-01', contract_end_date='2025-12-31',
        )
        self.client.login(username='admin', password='Str0ngAdminPass!')
        response = self.client.post(
            reverse('edit_employee', args=[employee.pk]),
            self.payload(full_name='موظف قائم', employee_type='خدمة مدنية'),
        )

        self.assertEqual(response.status_code, 302)
        employee.refresh_from_db()
        self.assertIsNone(employee.contract_start_date)
        self.assertIsNone(employee.contract_end_date)

class ArchiveReasonTests(SecurityTestCase):
    """نقل الموظف للأرشيف يسجّل سببه، والاستعادة تُفرّغه."""

    def setUp(self):
        super().setUp()
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        self.url = reverse('delete_employee', args=[self.own_employee.pk])

    def test_reason_is_stored_with_the_employee(self):
        response = self.client.post(self.url, {'archive_reason': 'انتهاء العقد'})

        self.assertEqual(response.status_code, 302)
        self.own_employee.refresh_from_db()
        self.assertTrue(self.own_employee.is_deleted)
        self.assertEqual(self.own_employee.archive_reason, 'انتهاء العقد')

    def test_reason_reaches_the_activity_log(self):
        self.client.post(self.url, {'archive_reason': 'نقل إلى منشأة أخرى'})

        self.assertTrue(
            ActivityLog.objects.filter(description__contains='نقل إلى منشأة أخرى').exists()
        )

    def test_missing_reason_does_not_archive(self):
        """النافذة تفرض السبب، لكن الطلب قد يصل بدونها."""
        response = self.client.post(self.url, {})

        self.assertEqual(response.status_code, 302)
        self.own_employee.refresh_from_db()
        self.assertFalse(self.own_employee.is_deleted)

    def test_whitespace_only_reason_does_not_archive(self):
        response = self.client.post(self.url, {'archive_reason': '   \n  '})

        self.assertEqual(response.status_code, 302)
        self.own_employee.refresh_from_db()
        self.assertFalse(self.own_employee.is_deleted)

    def test_overlong_reason_is_rejected(self):
        response = self.client.post(self.url, {'archive_reason': 'ا' * 501})

        self.assertEqual(response.status_code, 302)
        self.own_employee.refresh_from_db()
        self.assertFalse(self.own_employee.is_deleted)

    def test_reason_whitespace_is_collapsed(self):
        self.client.post(self.url, {'archive_reason': '  انتهاء    العقد  '})

        self.own_employee.refresh_from_db()
        self.assertEqual(self.own_employee.archive_reason, 'انتهاء العقد')

    def test_restore_clears_the_reason(self):
        """وإلا نُسب سبب أرشفة قديم إلى أرشفة لاحقة."""
        self.client.post(self.url, {'archive_reason': 'استقالة'})
        self.client.post(reverse('restore_employee', args=[self.own_employee.pk]))

        self.own_employee.refresh_from_db()
        self.assertFalse(self.own_employee.is_deleted)
        self.assertIsNone(self.own_employee.archive_reason)

    def test_restore_keeps_the_reason_in_the_activity_log(self):
        self.client.post(self.url, {'archive_reason': 'استقالة'})
        self.client.post(reverse('restore_employee', args=[self.own_employee.pk]))

        self.assertTrue(
            ActivityLog.objects.filter(
                action_type='استعادة', description__contains='استقالة'
            ).exists()
        )

    def test_reason_shows_in_the_archive_page(self):
        self.client.post(self.url, {'archive_reason': 'انتهاء الابتعاث'})
        response = self.client.get(reverse('archived_employee_list'))

        self.assertContains(response, 'انتهاء الابتعاث')

    def test_branch_user_cannot_archive_another_workplace(self):
        """الصلاحية القائمة لا تتأثر بإضافة السبب."""
        response = self.client.post(
            reverse('delete_employee', args=[self.other_employee.pk]),
            {'archive_reason': 'أي سبب'},
        )

        self.assertEqual(response.status_code, 404)
        self.other_employee.refresh_from_db()
        self.assertFalse(self.other_employee.is_deleted)


class SuperuserRedirectTests(SecurityTestCase):
    """رفض صفحات مدير النظام لا يجوز أن ينتج حلقة تحويل.

    صفحة الدخول تعيد المستخدم المسجَّل دخوله فوراً إلى next، فتوجيهه
    إليها عند الرفض كان يرتدّ بلا نهاية بدل أن يخبره بشيء.
    """

    ADMIN_ONLY = ['reconcile_employees', 'reconcile_template', 'user_list', 'system_settings']

    def test_branch_user_lands_on_the_dashboard_not_a_loop(self):
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        for name in self.ADMIN_ONLY:
            with self.subTest(view=name):
                response = self.client.get(reverse(name), follow=True)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.redirect_chain[-1][0], reverse('dashboard'))

    def test_anonymous_user_still_goes_to_login(self):
        for name in self.ADMIN_ONLY:
            with self.subTest(view=name):
                response = self.client.get(reverse(name))
                self.assertEqual(response.status_code, 302)
                self.assertTrue(response['Location'].startswith(reverse('login')))


class MinimumAgeTests(SecurityTestCase):
    """تاريخ الميلاد اختياري، لكنه إن وُجد فالعمر أكبر من 20 عاماً."""

    def setUp(self):
        super().setUp()
        self.client.login(username='admin', password='Str0ngAdminPass!')

    def payload(self, **overrides):
        data = {
            'full_name': 'موظف عمر',
            'gender': 'M',
            'nationality': self.saudi.pk,
            'status': 'نشط',
            'is_classified': 'غير مصنف',
            'has_sub_specialty': 'لا',
            'current_workplace': self.hospital_a.pk,
        }
        data.update(overrides)
        return data

    def dob_for_age(self, years, days_offset=0):
        """تاريخ ميلاد يعطي هذا العمر بالضبط اليوم."""
        today = date.today()
        try:
            born = today.replace(year=today.year - years)
        except ValueError:                       # 29 فبراير
            born = today.replace(year=today.year - years, month=2, day=28)
        return (born + timedelta(days=days_offset)).isoformat()

    def post(self, dob):
        return self.client.post(reverse('add_employee'), self.payload(dob=dob))

    def test_exactly_twenty_one_is_accepted(self):
        response = self.post(self.dob_for_age(21))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Employee.objects.filter(full_name='موظف عمر').exists())

    def test_comfortably_older_is_accepted(self):
        response = self.post(self.dob_for_age(45))

        self.assertEqual(response.status_code, 302)

    def test_exactly_twenty_is_rejected(self):
        """المطلوب «أكبر من 20»، فالعشرون تماماً مرفوضة."""
        response = self.post(self.dob_for_age(20))

        self.assertEqual(response.status_code, 200)
        self.assertIn('dob', response.context['form'].errors)
        self.assertFalse(Employee.objects.filter(full_name='موظف عمر').exists())

    def test_one_day_short_of_twenty_one_is_rejected(self):
        """الحدّ يُحسب بالسنوات المكتملة، لا بفارق السنوات وحده."""
        response = self.post(self.dob_for_age(21, days_offset=1))

        self.assertEqual(response.status_code, 200)
        self.assertIn('dob', response.context['form'].errors)

    def test_child_is_rejected(self):
        response = self.post(self.dob_for_age(8))

        self.assertEqual(response.status_code, 200)
        self.assertIn('dob', response.context['form'].errors)

    def test_future_date_is_rejected(self):
        """عمر سالب يسقط في الشرط نفسه بلا حاجة إلى فحص منفصل."""
        response = self.post((date.today() + timedelta(days=30)).isoformat())

        self.assertEqual(response.status_code, 200)
        self.assertIn('dob', response.context['form'].errors)

    def test_blank_date_is_still_allowed(self):
        """الحقل اختياري في قاعدة البيانات، فلا يصير إلزامياً بهذا التحقق."""
        response = self.client.post(reverse('add_employee'), self.payload())

        self.assertEqual(response.status_code, 302)
        self.assertIsNone(Employee.objects.get(full_name='موظف عمر').dob)

    def test_editing_an_underage_record_is_also_rejected(self):
        employee = Employee.objects.create(
            full_name='موظف قائم', gender='M', current_workplace=self.hospital_a
        )
        response = self.client.post(
            reverse('edit_employee', args=[employee.pk]),
            self.payload(full_name='موظف قائم', dob=self.dob_for_age(19)),
        )

        self.assertEqual(response.status_code, 200)
        employee.refresh_from_db()
        self.assertIsNone(employee.dob)

    def test_date_picker_is_capped_at_the_limit(self):
        """إرشاد في المتصفح، والفحص الحقيقي على الخادم."""
        form = EmployeeForm(user=self.admin)

        self.assertEqual(
            form.fields['dob'].widget.attrs['max'], latest_acceptable_dob().isoformat()
        )


class ArabicDaysAgoTests(SimpleTestCase):
    """صيغة «منذ كم يوم» تتبع قواعد تمييز العدد في العربية."""

    def ago(self, days, hours=0):
        moment = timezone.now() - timedelta(days=days, hours=hours)
        return days_ago(moment)

    def test_today_and_yesterday_are_named_not_counted(self):
        self.assertEqual(self.ago(0), 'اليوم')
        self.assertEqual(self.ago(1), 'أمس')

    def test_dual_is_genitive_after_the_preposition(self):
        """بعد «منذ» يُجرّ المثنى: يومين لا يومان."""
        self.assertEqual(self.ago(2), 'منذ يومين')

    def test_three_to_ten_take_the_broken_plural(self):
        self.assertEqual(self.ago(3), 'منذ 3 أيام')
        self.assertEqual(self.ago(10), 'منذ 10 أيام')

    def test_eleven_to_ninety_nine_take_the_accusative_singular(self):
        self.assertEqual(self.ago(11), 'منذ 11 يوماً')
        self.assertEqual(self.ago(99), 'منذ 99 يوماً')

    def test_hundred_and_above_take_the_genitive_singular(self):
        self.assertEqual(self.ago(100), 'منذ 100 يوم')
        self.assertEqual(self.ago(365), 'منذ 365 يوم')

    def test_a_calendar_day_counts_not_twenty_four_hours(self):
        """حدث قبل 20 ساعة قد يكون أمس، والعبرة بتغيّر اليوم لا بالساعات."""
        now = timezone.localtime()
        late_yesterday = (now - timedelta(days=1)).replace(hour=23, minute=30)
        self.assertEqual(days_ago(late_yesterday), 'أمس')

    def test_future_and_empty_do_not_crash(self):
        self.assertEqual(days_ago(timezone.now() + timedelta(days=2)), 'تاريخ لاحق')
        self.assertEqual(days_ago(None), '')


class ClassificationNumberTests(SecurityTestCase):
    """رقم التصنيف يلزم للمصنَّف، ويُمحى مع تاريخه لغير المصنَّف."""

    def setUp(self):
        super().setUp()
        self.client.login(username='admin', password='Str0ngAdminPass!')

    def payload(self, **overrides):
        data = {
            'full_name': 'موظف تصنيف',
            'gender': 'M',
            'nationality': self.saudi.pk,
            'status': 'نشط',
            'has_sub_specialty': 'لا',
            'current_workplace': self.hospital_a.pk,
            'is_classified': 'غير مصنف',
        }
        data.update(overrides)
        return data

    def test_classified_saves_number_and_expiry(self):
        response = self.client.post(reverse('add_employee'), self.payload(
            is_classified='مصنف',
            classification_number='SCFHS-12345',
            classification_expiry_date='2030-01-01',
        ))

        self.assertEqual(response.status_code, 302)
        employee = Employee.objects.get(full_name='موظف تصنيف')
        self.assertEqual(employee.classification_number, 'SCFHS-12345')
        self.assertEqual(str(employee.classification_expiry_date), '2030-01-01')

    def test_classified_without_a_number_is_rejected(self):
        response = self.client.post(reverse('add_employee'), self.payload(
            is_classified='مصنف', classification_expiry_date='2030-01-01',
        ))

        self.assertEqual(response.status_code, 200)
        self.assertIn('classification_number', response.context['form'].errors)
        self.assertFalse(Employee.objects.filter(full_name='موظف تصنيف').exists())

    def test_classified_without_an_expiry_is_still_rejected(self):
        response = self.client.post(reverse('add_employee'), self.payload(
            is_classified='مصنف', classification_number='SCFHS-1',
        ))

        self.assertEqual(response.status_code, 200)
        self.assertIn('classification_expiry_date', response.context['form'].errors)

    def test_whitespace_only_number_is_rejected(self):
        response = self.client.post(reverse('add_employee'), self.payload(
            is_classified='مصنف',
            classification_number='   ',
            classification_expiry_date='2030-01-01',
        ))

        self.assertEqual(response.status_code, 200)
        self.assertIn('classification_number', response.context['form'].errors)

    def test_unclassified_needs_neither(self):
        response = self.client.post(reverse('add_employee'), self.payload())

        self.assertEqual(response.status_code, 302)
        employee = Employee.objects.get(full_name='موظف تصنيف')
        self.assertIsNone(employee.classification_number)
        self.assertIsNone(employee.classification_expiry_date)

    def test_unclassifying_clears_a_stored_number(self):
        """وإلا بقي رقم تصنيف قديم في سجل من أُلغي تصنيفه."""
        employee = Employee.objects.create(
            full_name='موظف قائم', gender='M', current_workplace=self.hospital_a,
            is_classified='مصنف', classification_number='SCFHS-9',
            classification_expiry_date='2030-01-01',
        )
        response = self.client.post(
            reverse('edit_employee', args=[employee.pk]),
            self.payload(full_name='موظف قائم', is_classified='غير مصنف'),
        )

        self.assertEqual(response.status_code, 302)
        employee.refresh_from_db()
        self.assertIsNone(employee.classification_number)
        self.assertIsNone(employee.classification_expiry_date)

    def test_number_is_trimmed(self):
        self.client.post(reverse('add_employee'), self.payload(
            is_classified='مصنف',
            classification_number='  SCFHS-77  ',
            classification_expiry_date='2030-01-01',
        ))

        self.assertEqual(
            Employee.objects.get(full_name='موظف تصنيف').classification_number,
            'SCFHS-77',
        )


class FormTemplateIntegrityTests(SecurityTestCase):
    """يمنع تكرار الحقول في القالب: QueryDict يقرأ آخر قيمة، فيفوز الحقل الفارغ."""

    UNIQUE_FIELDS = (
        'classification_number', 'classification_expiry_date', 'is_classified',
        'contract_start_date', 'contract_end_date', 'employee_type',
        'full_name', 'national_id', 'date_of_birth',
    )

    def setUp(self):
        super().setUp()
        self.client.login(username='admin', password='Str0ngAdminPass!')

    def _assert_no_duplicates(self, url):
        html = self.client.get(url).content.decode()
        for field in self.UNIQUE_FIELDS:
            with self.subTest(field=field):
                self.assertLessEqual(
                    len(re.findall(r'name=["\']%s["\']' % re.escape(field), html)), 1,
                    'الحقل %s مكرر في %s' % (field, url),
                )

    def _assert_unique_ids(self, url):
        html = self.client.get(url).content.decode()
        ids = re.findall(r'\sid=["\']([^"\']+)["\']', html)
        duplicates = sorted({i for i in ids if ids.count(i) > 1})
        self.assertEqual(duplicates, [], 'معرّفات مكررة في %s: %s' % (url, duplicates))

    def test_add_form_has_no_duplicate_fields(self):
        self._assert_no_duplicates(reverse('add_employee'))

    def test_edit_form_has_no_duplicate_fields(self):
        self._assert_no_duplicates(reverse('edit_employee', args=[self.own_employee.pk]))

    def test_add_form_has_no_duplicate_ids(self):
        self._assert_unique_ids(reverse('add_employee'))

    def test_edit_form_has_no_duplicate_ids(self):
        self._assert_unique_ids(reverse('edit_employee', args=[self.own_employee.pk]))


class NationalitySettingsTests(SecurityTestCase):
    """الجنسيات صارت جدولاً مرجعياً يُدار من الإعدادات مثل المنشآت والأقسام."""

    def setUp(self):
        super().setUp()
        self.client.login(username='admin', password='Str0ngAdminPass!')

    def test_seed_list_is_present_and_ordered(self):
        """هجرة البيانات تزرع القائمة القديمة كاملة، بترتيبها لا أبجدياً."""
        self.assertGreaterEqual(Nationality.objects.count(), 190)
        for name in ('سعودي', 'مصري', 'كوت ديفوار', 'بدون جنسية', 'أخرى'):
            self.assertTrue(
                Nationality.objects.filter(name=name).exists(), f'{name} مفقودة'
            )
        self.assertEqual(Nationality.objects.first().name, 'سعودي')

    def test_appears_in_settings_page(self):
        response = self.client.get(reverse('system_settings'))
        self.assertContains(response, 'الجنسيات')
        self.assertContains(response, 'nationalitySearch')

    def test_admin_can_add_a_nationality(self):
        self.client.post(reverse('add_setting_item', args=['nationality']), {'name': 'مالطي جديد'})
        self.assertTrue(Nationality.objects.filter(name='مالطي جديد').exists())

    def test_new_nationality_goes_to_the_end_not_the_top(self):
        """وإلا قفزت الإضافات فوق «سعودي» في قائمة التسجيل."""
        self.client.post(reverse('add_setting_item', args=['nationality']), {'name': 'جنسية حديثة'})
        self.assertEqual(Nationality.objects.first().name, 'سعودي')
        self.assertEqual(Nationality.objects.last().name, 'جنسية حديثة')

    def test_branch_user_cannot_add_a_nationality(self):
        self.client.logout()
        self.client.login(username='branch_a', password='Str0ngPass!2024')
        self.client.post(reverse('add_setting_item', args=['nationality']), {'name': 'جنسية مهرَّبة'})
        self.assertFalse(Nationality.objects.filter(name='جنسية مهرَّبة').exists())

    def test_unused_nationality_can_be_deleted(self):
        spare = Nationality.objects.create(name='جنسية غير مستخدمة')
        self.client.post(reverse('delete_setting_item', args=['nationality', spare.pk]))
        self.assertFalse(Nationality.objects.filter(pk=spare.pk).exists())

    def test_nationality_in_use_is_protected(self):
        """PROTECT لا SET_NULL: الحذف يُرفض بدل أن يفرّغ الحقل في سجلات قائمة."""
        used = Nationality.objects.create(name='جنسية مستخدمة')
        employee = Employee.objects.create(
            full_name='موظف مرتبط', gender='M',
            current_workplace=self.hospital_a, nationality=used,
        )

        response = self.client.post(
            reverse('delete_setting_item', args=['nationality', used.pk]), follow=True
        )

        self.assertTrue(Nationality.objects.filter(pk=used.pk).exists())
        employee.refresh_from_db()
        self.assertEqual(employee.nationality, used)
        self.assertContains(response, 'مرتبط ببيانات موظفين')

    def test_renaming_reaches_every_linked_employee(self):
        """الفائدة الأساسية من الجدول: التصحيح الإملائي يسري على كل السجلات."""
        nationality = Nationality.objects.create(name='جنسيه بخطأ')
        employee = Employee.objects.create(
            full_name='موظف', gender='F',
            current_workplace=self.hospital_a, nationality=nationality,
        )

        nationality.name = 'جنسية صحيحة'
        nationality.save()

        employee.refresh_from_db()
        self.assertEqual(employee.nationality.name, 'جنسية صحيحة')

    def test_employee_form_saves_the_selected_nationality(self):
        egyptian = Nationality.objects.get(name='مصري')
        self.client.post(reverse('add_employee'), {
            'full_name': 'موظف جنسية', 'gender': 'M', 'status': 'نشط',
            'nationality': egyptian.pk, 'has_sub_specialty': 'لا',
            'is_classified': 'غير مصنف', 'employee_type': 'خدمة مدنية',
            'current_workplace': self.hospital_a.pk,
            'dob': (date.today() - timedelta(days=365 * 30)).isoformat(),
        })
        self.assertEqual(
            Employee.objects.get(full_name='موظف جنسية').nationality, egyptian
        )

    def test_export_writes_the_nationality_name(self):
        """الحقل صار مفتاحاً خارجياً، فلا يجوز أن يخرج كرقم في ملف Excel."""
        import io

        import openpyxl

        employee = Employee.objects.get(pk=self.own_employee.pk)
        employee.nationality = Nationality.objects.get(name='مصري')
        employee.save()

        sheet = openpyxl.load_workbook(
            io.BytesIO(self.client.get(reverse('export_employees_excel')).content)
        ).active
        header = [cell.value for cell in sheet[1]]
        row = next(
            r for r in sheet.iter_rows(min_row=2, values_only=True)
            if r[header.index('اسم الموظف')] == employee.full_name
        )
        self.assertEqual(row[header.index('الجنسية')], 'مصري')
